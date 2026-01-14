import streamlit as st
import pandas as pd
import os
import sys
import time
import openpyxl
import altair as alt
from io import BytesIO
from datetime import datetime, timedelta
from sqlalchemy.orm import Session

# --- Path Setup ---
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from agent import VendorResponseAgent
from ingest import create_vector_db
from database import init_db, SessionLocal, User, Document, AnswerBank

# --- CONFIGURATION ---
DATA_DIR = "data"
AUDIT_LOG_FILE = "audit_log.csv"
os.makedirs(DATA_DIR, exist_ok=True)

# --- DATABASE INITIALIZATION ---
if "db_initialized" not in st.session_state:
    init_db()
    st.session_state.db_initialized = True

# --- HELPER FUNCTIONS ---

def get_db():
    return SessionLocal()

def log_action(user, action, details):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if not os.path.exists(AUDIT_LOG_FILE):
        with open(AUDIT_LOG_FILE, "w") as f:
            f.write("Timestamp,User,Action,Details\n")
    with open(AUDIT_LOG_FILE, "a") as f:
        f.write(f"{timestamp},{user},{action},{details}\n")

# --- DB OPERATIONS ---

def db_get_answer_bank(search_term=None):
    db = get_db()
    query = db.query(AnswerBank)
    if search_term:
        search = f"%{search_term}%"
        query = query.filter(
            (AnswerBank.question.like(search)) | 
            (AnswerBank.answer.like(search)) | 
            (AnswerBank.product.like(search))
        )
    results = query.all()
    db.close()
    return [{
        "question": r.question, "answer": r.answer, 
        "product": r.product, "subsidiary": r.subsidiary, 
        "verified_by": r.verified_by, "date": r.date_added
    } for r in results]

def db_save_answer(question, answer, user, product, subsidiary):
    db = get_db()
    exists = db.query(AnswerBank).filter(AnswerBank.question == question).first()
    if not exists:
        new_entry = AnswerBank(
            question=question, answer=answer, 
            product=product, subsidiary=subsidiary,
            verified_by=user, date_added=datetime.now().strftime("%Y-%m-%d")
        )
        db.add(new_entry)
        db.commit()
        db.close()
        return True
    db.close()
    return False

def db_save_document(filename, desc, review_date, uploader):
    db = get_db()
    doc = db.query(Document).filter(Document.filename == filename).first()
    if not doc:
        doc = Document(
            filename=filename, description=desc,
            upload_date=datetime.now().strftime("%Y-%m-%d"),
            review_date=str(review_date), uploaded_by=uploader
        )
        db.add(doc)
    else:
        doc.description = desc
        doc.review_date = str(review_date)
    db.commit()
    db.close()

def db_get_documents():
    db = get_db()
    docs = db.query(Document).all()
    db.close()
    return docs

def db_delete_document(filename):
    db = get_db()
    doc = db.query(Document).filter(Document.filename == filename).first()
    if doc:
        db.delete(doc)
        db.commit()
    db.close()
    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        os.remove(path)

# --- PAGE CONFIG ---
st.set_page_config(page_title="AuditFlow Enterprise", page_icon="🛡️", layout="wide", initial_sidebar_state="expanded")

# --- SESSION STATE ---
if "theme_mode" not in st.session_state: st.session_state.theme_mode = "Light Mode"
if "auth_stage" not in st.session_state: st.session_state.auth_stage = "login"
if "page_selection" not in st.session_state: st.session_state.page_selection = "Executive Dashboard"

# Sync Session User with Database
if "user_profile" not in st.session_state:
    db = get_db()
    default_email = "john.smith@auditflow.io"
    user = db.query(User).filter(User.email == default_email).first()
    
    if not user:
        user = User(
            email=default_email, first_name="John", last_name="Smith",
            title="Sr. Security Analyst", phone="555-0199", role="Administrator"
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    
    st.session_state.user_profile = {
        "id": user.id, "first_name": user.first_name, "last_name": user.last_name,
        "email": user.email, "title": user.title, "phone": user.phone, "role": user.role
    }
    db.close()

def navigate_to(page_name):
    st.session_state.page_selection = page_name
    st.rerun()

def get_theme_css(mode):
    base = """
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    div[data-testid="stPopoverBody"] > div { padding: 10px !important; }
    .role-badge { background-color: #E0F2F1; color: #00695C; padding: 2px 8px; border-radius: 12px; font-size: 12px; font-weight: 600; margin-top: 4px; margin-bottom: 8px; display: inline-block; }
    """
    sidebar = """
    section[data-testid="stSidebar"] button { background-color: #F3F4F6 !important; color: #111827 !important; font-weight: 600 !important; border: none !important; }
    section[data-testid="stSidebar"] button:hover { background-color: #E5E7EB !important; color: #000000 !important; }
    """
    if mode == "Dark Mode":
        return base + sidebar + """section[data-testid="stSidebar"] { background-color: #1f1f1f; } .stApp { background-color: #0E1117; color: #E0E0E0; } div[data-testid="stMetric"] { background-color: #262730; border: 1px solid #444; border-radius: 8px; padding: 15px; }"""
    else:
        return base + sidebar + """section[data-testid="stSidebar"] { background-color: #111827; color: white; } section[data-testid="stSidebar"] * { color: #E5E7EB !important; } .stApp { background-color: #F9FAFB; color: #111827; } div[data-testid="stMetric"] { background-color: #ffffff; border: 1px solid #E5E7EB; border-radius: 8px; padding: 15px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }"""

st.markdown(f"<style>{get_theme_css(st.session_state.theme_mode)}</style>", unsafe_allow_html=True)

# --- HEADER ---
def show_header(title):
    c1, c2 = st.columns([6, 1])
    with c1: st.markdown(f"# {title}")
    with c2:
        u = st.session_state.user_profile
        initials = f"{u['first_name'][0]}{u['last_name'][0]}"
        with st.popover(f"👤 {initials}", use_container_width=True):
            st.markdown(f"**{u['first_name']} {u['last_name']}**")
            st.markdown(f"<span class='role-badge'>{u['role']}</span>", unsafe_allow_html=True)
            st.caption(u['title'])
            st.markdown("---")
            if st.button("⚙️ Manage Profile", use_container_width=True): navigate_to("Settings")
            if st.button("Log Out", key="logout_top", use_container_width=True):
                st.session_state.auth_stage = "login"
                st.rerun()

# --- AUTH FLOW ---
if st.session_state.auth_stage == "login":
    _, c, _ = st.columns([1,1,1])
    with c:
        st.markdown("<br><br>## AuditFlow Secure Login", unsafe_allow_html=True)
        st.info("Identity Provider: Azure AD")
        with st.form("login"):
            st.text_input("Username", value="john.smith@auditflow.io")
            st.text_input("Password", type="password", value="123")
            if st.form_submit_button("Sign In", type="primary", use_container_width=True):
                st.session_state.auth_stage = "mfa"
                st.rerun()
    st.stop()
elif st.session_state.auth_stage == "mfa":
    _, c, _ = st.columns([1,1,1])
    with c:
        st.markdown("<br><br>## MFA Challenge", unsafe_allow_html=True)
        st.warning("Enter code sent to +1 (555) ***-0199")
        if st.button("Verify 123456", type="primary", use_container_width=True):
            st.session_state.auth_stage = "authenticated"
            st.rerun()
    st.stop()

# --- SIDEBAR ---
with st.sidebar:
    st.title("AuditFlow")
    st.caption("Enterprise Compliance")
    st.markdown("---")
    pages = ["Executive Dashboard", "Auto-Fill (Beta)", "Answer Bank", "Gap Analysis", "My Projects", "Questionnaire Agent", "Knowledge Base", "Settings"]
    try: idx = pages.index(st.session_state.page_selection)
    except: idx = 0
    sel = st.selectbox("Navigation", pages, index=idx)
    if sel != st.session_state.page_selection:
        st.session_state.page_selection = sel
        st.rerun()
    st.markdown("---")
    st.caption(f"🟢 User: {st.session_state.user_profile['last_name']}")

# --- INITIALIZATION ---
if "agent" not in st.session_state: st.session_state.agent = VendorResponseAgent()
if "messages" not in st.session_state: st.session_state.messages = []

# --- PAGE 1: DASHBOARD ---
if st.session_state.page_selection == "Executive Dashboard":
    show_header("Executive Dashboard")
    docs = db_get_documents()
    bank = db_get_answer_bank()
    
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.metric("Security Posture", "Secure", "No Critical Flags")
    with c2: st.metric("Active Audits", "3", "+1 This Month")
    with c3: st.metric("Indexed Docs", len(docs), "Live")
    with c4: st.metric("Answer Bank", len(bank), "Verified Q&A")

    st.markdown("### Quick Actions")
    q1, q2, q3, q4 = st.columns(4)
    with q1: 
        if st.button("📄 Start Auto-Fill", use_container_width=True): navigate_to("Auto-Fill (Beta)")
    with q2:
        if st.button("📤 Upload Policy", use_container_width=True): navigate_to("Knowledge Base")
    with q3:
        if st.button("🔍 Run Gap Analysis", use_container_width=True): navigate_to("Gap Analysis")
    with q4:
        if st.button("🧠 Search Knowledge", use_container_width=True): navigate_to("Questionnaire Agent")

    st.divider()
    c_left, c_right = st.columns([2,1])
    with c_left:
        st.subheader("Audit Readiness")
        chart_data = pd.DataFrame({'Status': ['Completed', 'In Review', 'Drafting', 'Not Started'], 'Items': [85, 12, 15, 8]})
        st.altair_chart(alt.Chart(chart_data).mark_bar().encode(x='Items', y=alt.Y('Status', sort=None), color='Status').properties(height=250), use_container_width=True)
    with c_right:
        st.subheader("Recent Activity")
        if os.path.exists(AUDIT_LOG_FILE):
            st.dataframe(pd.read_csv(AUDIT_LOG_FILE).tail(5).sort_values(by="Timestamp", ascending=False), use_container_width=True, hide_index=True)

# --- PAGE 2: AUTO-FILL (TRUE EXCEL) ---
elif st.session_state.page_selection == "Auto-Fill (Beta)":
    show_header("Auto-Fill Assistant")
    up_file = st.file_uploader("Upload Excel Questionnaire", type=["xlsx"])
    if up_file:
        try:
            wb = openpyxl.load_workbook(up_file)
            sheet = st.selectbox("Select Sheet", wb.sheetnames)
            ws = wb[sheet]
            rows = list(ws.values)
            if len(rows) > 0:
                cols = [str(c) for c in rows[0]]
                st.markdown("#### Map Columns")
                q_col = st.selectbox("Question Column", cols)
                a_col = st.selectbox("Answer Column", cols)
                
                if st.button("🚀 Run Auto-Fill", type="primary"):
                    if not st.session_state.agent.vector_db: st.error("KB Empty!")
                    else:
                        prog = st.progress(0)
                        q_idx = cols.index(q_col)
                        a_idx = cols.index(a_col)
                        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
                            q_txt = str(row[q_idx].value) if row[q_idx].value else ""
                            if len(q_txt) > 5:
                                resp = st.session_state.agent.generate_responses([q_txt])
                                if not resp.empty:
                                    ws.cell(row=i, column=a_idx+1, value=resp.iloc[0]['AI_Response'])
                            prog.progress(min(i/ws.max_row, 1.0))
                        
                        out = BytesIO()
                        wb.save(out)
                        out.seek(0)
                        st.success("Done! Formatting Preserved.")
                        st.download_button("Download Result", out, f"filled_{up_file.name}", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        log_action("User", "AUTO_FILL", f"Processed {up_file.name}")
        except Exception as e: st.error(f"Error: {e}")

# --- PAGE 3: ANSWER BANK ---
elif st.session_state.page_selection == "Answer Bank":
    show_header("Answer Bank")
    c1, c2 = st.columns([4,1])
    with c1: search = st.text_input("Search...")
    with c2: 
        if st.button("➕ Add New", use_container_width=True): st.session_state.adding_new = True
    
    bank_data = db_get_answer_bank(search)
    if bank_data:
        st.dataframe(bank_data, use_container_width=True)
    else: st.info("No entries found.")

    if st.session_state.get("adding_new"):
        st.divider()
        with st.form("new_ans"):
            st.write("Add Verified Answer")
            c_a, c_b = st.columns(2)
            with c_a: p = st.text_input("Product")
            with c_b: s = st.text_input("Subsidiary")
            q = st.text_input("Question")
            a = st.text_area("Answer")
            if st.form_submit_button("Save"):
                if db_save_answer(q, a, st.session_state.user_profile["last_name"], p, s):
                    st.success("Saved!")
                    st.session_state.adding_new = False
                    st.rerun()
                else: st.warning("Duplicate question.")

# --- PAGE 4: GAP ANALYSIS (FULL RESTORED) ---
elif st.session_state.page_selection == "Gap Analysis":
    show_header("Strategic Gap Analysis")
    st.markdown("Scan your knowledge base against common compliance frameworks.")
    
    framework = st.selectbox("Select Framework", ["SOC 2 Type II", "ISO 27001:2022", "NIST 800-53"])
    
    if st.button("🔍 Run Gap Analysis", type="primary"):
        with st.spinner(f"Scanning documents against {framework} controls..."):
            time.sleep(1.5) # Simulation
            
            col1, col2, col3 = st.columns(3)
            with col1: st.metric("Controls Covered", "85%", "+5% vs Last Scan")
            with col2: st.metric("Missing Policies", "3", "Critical")
            with col3: st.metric("Evidence Strength", "Medium", "Needs Improvement")
            
            st.divider()
            st.subheader("⚠️ Missing or Weak Controls")
            
            issues = [
                {"Control": "CC-6.1", "Area": "Vulnerability Management", "Status": "Missing", "Suggestion": "Upload a 'Vulnerability Scanning Policy'"},
                {"Control": "CC-8.1", "Area": "Change Management", "Status": "Partial", "Suggestion": "Current 'DevOps Guide' lacks rollback procedures."},
                {"Control": "A.12.3", "Area": "Backup", "Status": "Verified", "Suggestion": "None. 'Backup_Policy_2025.pdf' covers this."}
            ]
            st.dataframe(pd.DataFrame(issues), use_container_width=True, hide_index=True)

# --- PAGE 5: ACTIVE PROJECTS (FULL RESTORED) ---
elif st.session_state.page_selection == "My Projects":
    show_header("Active Questionnaires")
    st.info("Select a project below to view details and manage status.")
    
    projects = pd.DataFrame({
        "Project Name": ["SoundThinking SIG 2026", "Internal ISO Audit", "Vendor A - CAIQ Lite"],
        "Due Date": ["Feb 28, 2026", "Mar 15, 2026", "Jan 10, 2026"],
        "Progress": [65, 20, 90],
        "Type": ["SIG Core", "ISO 27001", "CAIQ"]
    })
    
    event = st.dataframe(projects, use_container_width=True, hide_index=True, selection_mode="single-row", on_select="rerun", column_config={"Progress": st.column_config.ProgressColumn("Completion", format="%d%%", min_value=0, max_value=100)})
    
    if len(event.selection.rows) > 0:
        selected_index = event.selection.rows[0]
        selected_project = projects.iloc[selected_index]
        st.divider()
        st.subheader(f"📂 Managing: {selected_project['Project Name']}")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"**Due Date:** {selected_project['Due Date']}")
            st.button("📅 Change Deadline", key="btn_date")
        with c2:
            st.markdown(f"**Type:** {selected_project['Type']}")
            st.button("📤 Export Draft", key="btn_export")
        with c3:
            st.markdown(f"**Status:** {selected_project['Progress']}% Complete")
            if st.button("✅ Mark Complete", key="btn_complete", type="primary"):
                st.balloons()
                st.success("Project marked as complete!")

# --- PAGE 6: AGENT (FULL RESTORED) ---
elif st.session_state.page_selection == "Questionnaire Agent":
    show_header("Vendor Response Agent")
    with st.expander("ℹ️ How to use this Agent"):
        st.markdown("1. **Ask a question:** Type naturally.\n2. **Review Evidence:** Click 'Verified Source'.\n3. **Save to Answer Bank:** Add good answers to the memory.")

    if len(st.session_state.messages) > 0:
        col_export, _ = st.columns([1, 5])
        with col_export:
            export_data = [{"Role": m["role"], "Content": m["content"], "Evidence": m.get("evidence", "")} for m in st.session_state.messages]
            st.download_button(label="📥 Download Report", data=pd.DataFrame(export_data).to_csv(index=False).encode('utf-8'), file_name="audit_report.csv", mime="text/csv")

    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
            if message.get("evidence"): 
                with st.expander("🔍 Source"): st.markdown(message["evidence"])

    if prompt := st.chat_input("Ask a question..."):
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"): st.markdown(prompt)
        with st.chat_message("assistant"):
            with st.spinner("Thinking..."):
                try:
                    df = st.session_state.agent.generate_responses([prompt])
                    if not df.empty:
                        answer, evidence = df.iloc[0]['AI_Response'], df.iloc[0]['Evidence']
                        st.markdown(answer)
                        if evidence and evidence != "No Source": 
                            with st.expander("🔍 Source"): st.markdown(evidence)
                            if st.button("💾 Save to Bank"):
                                save_to_answer_bank(prompt, answer, st.session_state.user_profile["last_name"], "General", "All")
                                st.success("Saved!")
                        st.session_state.messages.append({"role": "assistant", "content": answer, "evidence": evidence})
                        log_action("User", "QUERY_AI", prompt[:50] + "...")
                    else: st.error("No response.")
                except Exception as e: st.error(f"Error: {e}")

# --- PAGE 7: KNOWLEDGE BASE ---
elif st.session_state.page_selection == "Knowledge Base":
    show_header("Knowledge Base")
    with st.expander("Upload Documents"):
        up_files = st.file_uploader("Select Files", accept_multiple_files=True)
        if up_files:
            r_date = st.date_input("Review Date", value=datetime.now() + timedelta(days=365))
            desc_map = {f.name: st.text_input(f"Desc: {f.name}") for f in up_files}
            if st.button("Process"):
                os.makedirs("data", exist_ok=True)
                for f in up_files:
                    with open(os.path.join("data", f.name), "wb") as w: w.write(f.getbuffer())
                    db_save_document(f.name, desc_map[f.name], r_date, st.session_state.user_profile["last_name"])
                create_vector_db()
                st.session_state.agent = VendorResponseAgent()
                st.rerun()
    
    st.divider()
    docs = db_get_documents()
    if docs:
        for d in docs:
            c1, c2, c3 = st.columns([3,2,1])
            with c1: st.write(f"**{d.filename}**\n<small>{d.description}</small>", unsafe_allow_html=True)
            with c2: st.caption(f"Review: {d.review_date}")
            with c3:
                if st.button("🗑️", key=f"del_{d.id}"):
                    db_delete_document(d.filename)
                    create_vector_db()
                    st.rerun()
    else: st.info("No documents.")

# --- PAGE 8: SETTINGS (FULL RESTORED) ---
elif st.session_state.page_selection == "Settings":
    show_header("Settings")
    t1, t2, t3, t4 = st.tabs(["Profile", "Roles & Permissions", "Theme", "Logs"])
    
    with t1:
        u = st.session_state.user_profile
        c1, c2 = st.columns(2)
        with c1:
            fn = st.text_input("First Name", u['first_name'])
            ln = st.text_input("Last Name", u['last_name'])
            em = st.text_input("Email", u['email'])
        with c2:
            ti = st.text_input("Title", u['title'])
            ph = st.text_input("Phone", u['phone'])
            # Explicit Role Toggle
            rl = st.selectbox("Role", ["Administrator", "Manager", "Analyst", "Read Only"], index=["Administrator", "Manager", "Analyst", "Read Only"].index(u['role']))
        
        if st.button("Save Profile"):
            # Update Session
            st.session_state.user_profile.update({"first_name": fn, "last_name": ln, "email": em, "title": ti, "phone": ph, "role": rl})
            # Update DB
            db = get_db()
            user_rec = db.query(User).filter(User.id == u['id']).first()
            if user_rec:
                user_rec.first_name = fn; user_rec.last_name = ln; user_rec.email = em
                user_rec.title = ti; user_rec.phone = ph; user_rec.role = rl
                db.commit()
            db.close()
            st.success("Updated!")
            st.rerun()

    with t2:
        st.markdown("### Role Management")
        st.caption("Manage access levels for the organization.")
        if "role_data" not in st.session_state:
            st.session_state.role_data = pd.DataFrame({"Role": ["Administrator", "Analyst", "Auditor", "Viewer"], "Write Access": [True, True, False, False], "Delete Access": [True, False, False, False], "AI Access": [True, True, True, False]})
        edited_df = st.data_editor(st.session_state.role_data, num_rows="dynamic", use_container_width=True, key="role_editor")
        if st.button("💾 Save Permission Changes", type="primary"):
            st.session_state.role_data = edited_df
            st.success("Permissions updated successfully!")
            log_action("Admin", "UPDATE_ROLES", "Modified system role permissions")

    with t3:
        st.radio("Theme", ["Pro (Default)", "Dark Mode", "Light Mode"], key="theme_sel")
        if st.session_state.theme_sel != st.session_state.theme_mode:
            st.session_state.theme_mode = st.session_state.theme_sel
            st.rerun()

    with t4:
        st.markdown("### System Audit Logs")
        if os.path.exists(AUDIT_LOG_FILE): st.dataframe(pd.read_csv(AUDIT_LOG_FILE).sort_values(by="Timestamp", ascending=False), use_container_width=True)
